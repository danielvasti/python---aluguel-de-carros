import openpyxl

# Carregar o arquivo Excel
arquivo = "carros.xlsx"

# Função para carregar os dados de uma aba
def carregar_dados(aba_nome):
    wb = openpyxl.load_workbook(arquivo)
    aba = wb[aba_nome]
    dados = []
    for row in aba.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            dados.append(row)
    wb.close()
    return dados

# Função para salvar dados em uma aba
def salvar_dados(aba_nome, dados):
    wb = openpyxl.load_workbook(arquivo)
    aba = wb[aba_nome]
    aba.delete_rows(2, aba.max_row)
    for row in dados:
        aba.append(row)
    wb.save(arquivo)
    wb.close()

# Função para listar carros disponíveis
def listar_carros_disponiveis():
    carros = carregar_dados("carros_disponiveis")
    if not carros:
        print("Nenhum carro disponível.")
    else:
        print("Carros disponíveis para aluguel:")
        for carro in carros:
            print(f"{carro[0]}. {carro[1]} - R$ {carro[2]} por dia")

# Função para listar carros alugados
def listar_carros_alugados():
    carros = carregar_dados("carros_alugados")
    if not carros:
        print("Nenhum carro alugado no momento.")
    else:
        print("Carros alugados:")
        for carro in carros:
            print(f"{carro[0]}. {carro[1]} - R$ {carro[2]} por dia")

# Função para alugar um carro
def alugar_carro(carro_id, dias):
    carros_disponiveis = carregar_dados("carros_disponiveis")
    carros_alugados = carregar_dados("carros_alugados")
    carro_encontrado = None
    for carro in carros_disponiveis:
        if carro[0] == carro_id:
            carro_encontrado = carro
            break
    if carro_encontrado:
        total = carro_encontrado[2] * dias
        print(f"Você alugou o carro {carro_encontrado[1]} por {dias} dias. Total: R$ {total}")
        carros_disponiveis.remove(carro_encontrado)
        carros_alugados.append(carro_encontrado)
        salvar_dados("carros_disponiveis", carros_disponiveis)
        salvar_dados("carros_alugados", carros_alugados)
    else:
        print("Carro não encontrado!")

# Função para devolver um carro
def devolver_carro(carro_id):
    carros_disponiveis = carregar_dados("carros_disponiveis")
    carros_alugados = carregar_dados("carros_alugados")
    carro_encontrado = None
    for carro in carros_alugados:
        if carro[0] == carro_id:
            carro_encontrado = carro
            break
    if carro_encontrado:
        print(f"Você devolveu o carro {carro_encontrado[1]}.")
        carros_alugados.remove(carro_encontrado)
        carros_disponiveis.append(carro_encontrado)
        salvar_dados("carros_disponiveis", carros_disponiveis)
        salvar_dados("carros_alugados", carros_alugados)
    else:
        print("Carro não encontrado!")

# Interface principal
while True:
    print("\nBem vindo à locadora de carros!")
    print("1 - Ver carros disponíveis")
    print("2 - Alugar um carro")
    print("3 - Devolver um carro")
    print("4 - Ver carros alugados")
    print("5 - Sair")

    opcao = input("Escolha uma opção: ")

    if opcao == "1":
        listar_carros_disponiveis()

    elif opcao == "2":
        listar_carros_disponiveis()
        carro_id = int(input("Digite o ID do carro que deseja alugar: "))
        dias = int(input("Por quantos dias você deseja alugar o carro? "))
        alugar_carro(carro_id, dias)

    elif opcao == "3":
        listar_carros_alugados()
        carro_id = int(input("Digite o ID do carro que deseja devolver: "))
        devolver_carro(carro_id)

    elif opcao == "4":
        listar_carros_alugados()

    elif opcao == "5":
        print("Obrigado por usar nossos serviços!")
        break

    else:
        print("Opção inválida! Tente novamente.")
